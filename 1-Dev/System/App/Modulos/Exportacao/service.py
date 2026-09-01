from __future__ import annotations

import hashlib
import html
import json
import os
import signal
import shutil
import subprocess
import tempfile
import time
from copy import copy, deepcopy
from datetime import datetime
from pathlib import Path
from urllib.parse import quote
from zoneinfo import ZoneInfo

from Core.atomic import atomic_write_json
from Core.config import runtime_component_installation_root, runtime_libreoffice, runtime_libreoffice_candidates, station_id
from Core.resources import resource_target, verify_installed_resource
from Core.context import current_identity
from Core.db import connect
from Core.ids import next_id
from Core.filetx import current as current_file_transaction, stage_directory
from Core.paths import DATA_ROOT, ensure_under_data
from Core.storage import ensure_data_quota
from Core.version import app_version
from Core.provenance import public_notice
from Core.release import verify_runtime_component
from Modulos.Carregamentos import service as carregamentos


TEMPLATE_PATH = (
    Path(__file__).resolve().parents[2]
    / "Templates"
    / "Exportacao"
    / "TEMPLATE_PADRAO_CARREGAMENTO_FINAL.xlsx"
)
BASE_WORK_ROWS = 6
BASE_COST_ROWS = 7
BASE_ITEM_ROWS = 18
CURRENCY_FORMAT = 'R$ #,##0.00'
PERCENT_FORMAT = '0.00%'


def _now() -> str:
    timezone = os.environ.get("CJL_TIMEZONE", "America/Sao_Paulo")
    return datetime.now(ZoneInfo(timezone)).isoformat(timespec="seconds")


def _visual_datetime(value: str) -> str:
    try:
        return datetime.fromisoformat(str(value)).strftime("%d/%m/%Y - %H:%M")
    except Exception:
        return str(value or "")


def _sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _sha256_file(path: Path) -> str:
    return _sha256_bytes(path.read_bytes())


def _json_hash(value) -> str:
    raw = json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str, allow_nan=False
    ).encode("utf-8")
    return _sha256_bytes(raw)


def _document_content(load: dict) -> dict:
    ignored = {"documentos", "pode_editar", "bloqueado"}
    return {key: value for key, value in load.items() if key not in ignored}


def _excel_safe(value):
    if isinstance(value, dict):
        return {key: _excel_safe(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_excel_safe(item) for item in value]
    if isinstance(value, tuple):
        return tuple(_excel_safe(item) for item in value)
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _safe_output(path: Path) -> Path:
    return ensure_under_data(path)


def _document_root(load: dict, revision: int) -> Path:
    generated = datetime.fromisoformat(str(load.get("data")))
    return _safe_output(
        DATA_ROOT
        / "Documentos"
        / str(generated.year)
        / f"{generated.month:02d}"
        / load["id"]
        / f"REV_{revision:03d}"
    )


def _require_openpyxl():
    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image
    except ImportError as exc:
        raise RuntimeError(
            "O RUNTIME PYTHON ASSINADO ESTÁ INCOMPLETO: OPENPYXL NÃO FOI ENCONTRADO EM "
            "SISTEMA\\RUNTIME\\PYTHON. RESTAURE UM PACOTE OFICIAL."
        ) from exc
    return load_workbook, Image


def _work_items(load: dict, work: dict) -> list[dict]:
    return [row for row in load.get("itens", []) if row.get("obra_id") == work["id"]]


def _copy_sheet(source, target) -> None:
    for row in source.iter_rows():
        for cell in row:
            destination = target[cell.coordinate]
            destination.value = cell.value
            if cell.has_style:
                destination._style = copy(cell._style)
            destination.number_format = cell.number_format
            destination.alignment = copy(cell.alignment)
            destination.protection = copy(cell.protection)
    for key, dimension in source.column_dimensions.items():
        target.column_dimensions[key].width = dimension.width
        target.column_dimensions[key].hidden = dimension.hidden
    for key, dimension in source.row_dimensions.items():
        target.row_dimensions[key].height = dimension.height
        target.row_dimensions[key].hidden = dimension.hidden
    for merged in source.merged_cells.ranges:
        target.merge_cells(str(merged))
    target.sheet_view.showGridLines = source.sheet_view.showGridLines
    target.freeze_panes = source.freeze_panes
    target.page_margins = copy(source.page_margins)
    target.page_setup = copy(source.page_setup)
    target.print_options = copy(source.print_options)
    target.sheet_properties = copy(source.sheet_properties)


def _insert_rows_preserving_merges(sheet, index: int, amount: int) -> None:
    if amount <= 0:
        return
    shifted = []
    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row < index <= merged.max_row:
            raise RuntimeError(f"O TEMPLATE POSSUI UMA MESCLAGEM QUE ATRAVESSA A LINHA {index}.")
        if merged.min_row >= index:
            shifted.append(
                (
                    merged.min_row + amount,
                    merged.min_col,
                    merged.max_row + amount,
                    merged.max_col,
                )
            )
            sheet.unmerge_cells(str(merged))
    sheet.insert_rows(index, amount)
    for min_row, min_col, max_row, max_col in shifted:
        sheet.merge_cells(
            start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col
        )


def _copy_row_style(sheet, source_row: int, destination_row: int, first_col: int, last_col: int) -> None:
    sheet.row_dimensions[destination_row].height = sheet.row_dimensions[source_row].height
    for column in range(first_col, last_col + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(destination_row, column)
        target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def _expand_work_sheet_items(sheet, required: int) -> tuple[int, int, int]:
    extra = max(0, required - BASE_ITEM_ROWS)
    _insert_rows_preserving_merges(sheet, 32, extra)
    for row in range(32, 32 + extra):
        _copy_row_style(sheet, 14, row, 2, 12)
        sheet.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
        sheet.merge_cells(start_row=row, start_column=11, end_row=row, end_column=12)
    return 14, 31 + extra, extra


def _place_item_image(sheet, item: dict, row: int, Image) -> None:
    name = str(item.get("imagem_arquivo") or "").strip()
    if not name:
        sheet.cell(row, 3).value = "SEM IMAGEM"
        return
    source = _safe_output(DATA_ROOT / "Imagens" / "Equipamentos" / name)
    if not source.is_file():
        sheet.cell(row, 3).value = "IMAGEM INDISPONÍVEL"
        return
    picture = Image(str(source))
    column_width = float(sheet.column_dimensions["C"].width or 10.75)
    cell_width_px = max(24.0, column_width * 7.0 + 5.0)
    row_points = max(float(sheet.row_dimensions[row].height or 56.25), 56.25)
    cell_height_px = row_points * 96.0 / 72.0
    scale = min((cell_width_px * 0.90) / max(float(picture.width), 1.0), (cell_height_px * 0.90) / max(float(picture.height), 1.0))
    picture.width = max(1, int(float(picture.width) * scale))
    picture.height = max(1, int(float(picture.height) * scale))
    picture.anchor = f"C{row}"
    sheet.add_image(picture)
    sheet.row_dimensions[row].height = row_points


def _fill_work_sheet(
    sheet,
    load: dict,
    work: dict,
    items: list[dict],
    generated_at: str,
    content_hash: str,
    Image,
) -> None:
    sheet["J3"] = work.get("op_numero") or "—"
    sheet["K4"] = load.get("revisao_operacional") or 1
    sheet["L4"] = _visual_datetime(generated_at)
    sheet["D7"] = f"{work.get('estado') or ''} / {work.get('uf') or ''}".strip(" / ")
    sheet["I7"] = work.get("municipio") or "—"
    sheet["D8"] = work.get("nome") or "—"
    sheet["I8"] = load["id"]
    contract = " · ".join(
        value
        for value in (work.get("cliente_nome"), work.get("referencia_contrato"))
        if value
    )
    sheet["D9"] = contract or "—"
    sheet["I9"] = work.get("previsao_entrega") or "—"
    geo = (
        f"{work.get('latitude')}, {work.get('longitude')}"
        if work.get("latitude") is not None and work.get("longitude") is not None
        else work.get("endereco")
    )
    sheet["D10"] = geo or "NÃO INFORMADA"
    if work.get("latitude") is not None and work.get("longitude") is not None:
        sheet["D10"].hyperlink = f"https://www.google.com/maps?q={work.get('latitude')},{work.get('longitude')}"
        sheet["D10"].style = "Hyperlink"
    elif work.get("endereco"):
        sheet["D10"].hyperlink = f"https://www.google.com/maps/search/?api=1&query={quote(str(work.get('endereco')), safe='')}"
        sheet["D10"].style = "Hyperlink"
    sheet["I10"] = float(work.get("valor_obra") or 0)
    sheet["I10"].number_format = CURRENCY_FORMAT
    first_row, last_row, extra = _expand_work_sheet_items(sheet, len(items))
    for index, row in enumerate(range(first_row, last_row + 1)):
        item = items[index] if index < len(items) else None
        sheet.cell(row, 2).value = index + 1 if item else None
        sheet.cell(row, 3).value = None
        sheet.cell(row, 4).value = (
            f"{item.get('equipamento_codigo')} · {item.get('equipamento_nome')}" if item else None
        )
        sheet.cell(row, 9).value = int(item.get("quantidade") or 0) if item else None
        sheet.cell(row, 10).value = (item.get("unidade") or "UN") if item else None
        sheet.cell(row, 11).value = (item.get("observacao") or "") if item else None
        if item:
            _place_item_image(sheet, item, row, Image)
    summary_row = 33 + extra
    sheet.cell(summary_row, 9).value = sum(int(item.get("quantidade") or 0) for item in items)
    sheet.cell(summary_row, 11).value = f"{len(items)} tipo(s)"
    sheet.cell(summary_row + 3, 2).value = work.get("observacao") or "SEM OBSERVAÇÕES."
    sheet.cell(summary_row + 8, 7).value = load.get("solicitante") or "NÃO INFORMADO"
    sheet.cell(summary_row + 9, 7).value = content_hash
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_area = f"B2:L{42 + extra}"


def _expand_cost_sheet(sheet, works: int, personnel: int, freight: int) -> tuple[int, int, int]:
    work_extra = max(0, works - BASE_WORK_ROWS)
    _insert_rows_preserving_merges(sheet, 21, work_extra)
    for row in range(21, 21 + work_extra):
        _copy_row_style(sheet, 20, row, 2, 14)
        for start, end in ((4, 6), (7, 8), (9, 10), (12, 13)):
            sheet.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
    cost_start = 24 + work_extra
    cost_extra = max(0, max(personnel, freight) - BASE_COST_ROWS)
    _insert_rows_preserving_merges(sheet, cost_start + BASE_COST_ROWS, cost_extra)
    for row in range(cost_start + BASE_COST_ROWS, cost_start + BASE_COST_ROWS + cost_extra):
        _copy_row_style(sheet, cost_start + BASE_COST_ROWS - 1, row, 2, 14)
        sheet.merge_cells(start_row=row, start_column=8, end_row=row, end_column=10)
        sheet.merge_cells(start_row=row, start_column=11, end_row=row, end_column=12)
    return work_extra, cost_extra, cost_start


def _write_cost_sheet(sheet, load: dict, generated_at: str) -> tuple[int, int]:
    works = load.get("obras", [])
    personnel = [row for row in load.get("custos", []) if row.get("grupo") == "PESSOAL"]
    freight = [row for row in load.get("custos", []) if row.get("grupo") == "FRETE"]
    work_extra, cost_extra, cost_start = _expand_cost_sheet(
        sheet, len(works), len(personnel), len(freight)
    )
    sheet["L3"] = load["id"]
    sheet["M4"] = load.get("revisao_operacional") or 1
    sheet["N4"] = _visual_datetime(generated_at)
    sheet["D7"] = "FÁBRICA / ORIGEM"
    sheet["I7"] = load.get("data_saida") or load.get("data") or "—"
    sheet["M7"] = load.get("data_retorno") or "—"
    sheet["D8"] = load.get("veiculo") or "NÃO INFORMADO"
    sheet["I8"] = load.get("placa") or "SEM PLACA"
    sheet["M8"] = load.get("motorista") or "NÃO INFORMADO"
    sheet["D9"] = f"{int(load.get('funcionarios') or 0)} funcionário(s)"
    sheet["I9"] = float(load.get("dias_viagem") or 0)
    sheet["M9"] = f"{float(load.get('distancia_km') or 0):.2f} km"
    sheet["D10"] = " → ".join(f"{row.get('municipio')} / {row.get('uf')}" for row in works)
    sheet["M10"] = load.get("observacao") or "SEM OBSERVAÇÕES."
    reserved_works = max(BASE_WORK_ROWS, len(works))
    for index, row in enumerate(range(15, 15 + reserved_works)):
        work = works[index] if index < len(works) else None
        values = (
            [
                index + 1,
                work.get("op_numero"),
                work.get("nome"),
                f"{work.get('municipio')} / {work.get('uf')}",
                (
                    f"{work.get('latitude')}, {work.get('longitude')}"
                    if work.get("latitude") is not None
                    else work.get("endereco")
                ),
                float(work.get("valor_obra") or 0),
                work.get("previsao_entrega"),
                work.get("observacao"),
            ]
            if work
            else [None] * 8
        )
        for column, value in zip((2, 3, 4, 7, 9, 11, 12, 14), values):
            sheet.cell(row, column).value = value
        sheet.cell(row, 11).number_format = CURRENCY_FORMAT
    reserved_costs = max(BASE_COST_ROWS, len(personnel), len(freight))
    for index, row in enumerate(range(cost_start, cost_start + reserved_costs)):
        team = personnel[index] if index < len(personnel) else None
        sheet.cell(row, 2).value = team.get("descricao") if team else None
        sheet.cell(row, 3).value = float(team.get("valor_unitario") or 0) if team else 0
        sheet.cell(row, 4).value = int(team.get("funcionarios_aplicados") or 0) if team else 0
        sheet.cell(row, 5).value = float(team.get("quantidade") or 0) if team else 0
        sheet.cell(row, 6).value = float(team.get("total") or 0) if team else 0
        sheet.cell(row, 6).number_format = CURRENCY_FORMAT
        trip = freight[index] if index < len(freight) else None
        sheet.cell(row, 8).value = trip.get("descricao") if trip else None
        sheet.cell(row, 11).value = float(trip.get("valor_unitario") or 0) if trip else 0
        sheet.cell(row, 13).value = float(trip.get("quantidade") or 0) if trip else 0
        sheet.cell(row, 14).value = float(trip.get("total") or 0) if trip else 0
        sheet.cell(row, 14).number_format = CURRENCY_FORMAT
    total_row = cost_start + reserved_costs
    team_total = round(sum(float(row.get("total") or 0) for row in personnel), 2)
    freight_total = round(sum(float(row.get("total") or 0) for row in freight), 2)
    sheet.cell(total_row, 6).value = team_total
    sheet.cell(total_row, 14).value = freight_total
    sheet.cell(total_row, 6).number_format = CURRENCY_FORMAT
    sheet.cell(total_row, 14).number_format = CURRENCY_FORMAT
    summary_row = 34 + work_extra + cost_extra
    value_total = round(sum(float(work.get("valor_obra") or 0) for work in works), 2)
    values = {
        (summary_row, 7): value_total,
        (summary_row, 13): ((team_total + freight_total) / value_total if value_total else 0),
        (summary_row + 1, 7): team_total,
        (summary_row + 1, 13): freight_total,
        (summary_row + 2, 7): team_total + freight_total,
        (summary_row + 2, 13): ((team_total + freight_total) / len(works) if works else 0),
    }
    for (row, column), value in values.items():
        sheet.cell(row, column).value = value
        sheet.cell(row, column).number_format = (
            PERCENT_FORMAT if (row, column) == (summary_row, 13) else CURRENCY_FORMAT
        )
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_area = f"B2:N{40 + work_extra + cost_extra}"
    return work_extra, cost_extra


def _expand_summary_sheet(sheet, works: int) -> int:
    extra = max(0, works - 10)
    _insert_rows_preserving_merges(sheet, 24, extra)
    for row in range(24, 24 + extra):
        _copy_row_style(sheet, 23, row, 2, 13)
        sheet.merge_cells(start_row=row, start_column=12, end_row=row, end_column=13)
    return extra


def _write_summary_sheet(sheet, load: dict, generated_at: str) -> int:
    works = load.get("obras", [])
    extra = _expand_summary_sheet(sheet, len(works))
    sheet["K3"] = load["id"]
    sheet["K4"] = _visual_datetime(generated_at)
    sheet["D10"] = " → ".join(f"{row.get('municipio')} / {row.get('uf')}" for row in works)
    sheet["J10"] = (
        " · ".join(value for value in (load.get("veiculo"), load.get("placa")) if value)
        or "NÃO INFORMADO"
    )
    reserved_works = max(10, len(works))
    for index, row in enumerate(range(14, 14 + reserved_works)):
        work = works[index] if index < len(works) else None
        values = (
            [
                index + 1,
                work.get("op_numero"),
                work.get("nome"),
                f"{work.get('municipio')} / {work.get('uf')}",
                float(work.get("valor_obra") or 0),
                float(work.get("custo_frete") or 0),
                float(work.get("custo_pessoal") or 0),
                float(work.get("custo_viagem") or 0),
                (
                    float(work.get("custo_viagem") or 0) / float(work.get("valor_obra") or 1)
                ),
                float(work.get("valor_obra") or 0) - float(work.get("custo_viagem") or 0),
                work.get("observacao") or "",
            ]
            if work
            else [None] * 11
        )
        for column, value in enumerate(values, start=2):
            sheet.cell(row, column).value = value
        for column in (6, 7, 8, 9, 11):
            sheet.cell(row, column).number_format = CURRENCY_FORMAT
        sheet.cell(row, 10).number_format = PERCENT_FORMAT
    total_row = 24 + extra
    for column in range(6, 12):
        letter = sheet.cell(1, column).column_letter
        sheet.cell(total_row, column).value = (
            f"=IFERROR(I{total_row}/F{total_row},0)"
            if column == 10
            else f"=SUM({letter}14:{letter}{total_row - 1})"
        )
    for column in (6, 7, 8, 9, 11):
        sheet.cell(total_row, column).number_format = CURRENCY_FORMAT
    sheet.cell(total_row, 10).number_format = PERCENT_FORMAT
    team_total = round(sum(float(work.get("custo_pessoal") or 0) for work in works), 2)
    freight_total = round(sum(float(work.get("custo_frete") or 0) for work in works), 2)
    value_total = round(sum(float(work.get("valor_obra") or 0) for work in works), 2)
    for cell, value in (
        ("B7", value_total),
        ("E7", team_total),
        ("H7", freight_total),
        ("K7", team_total + freight_total),
    ):
        sheet[cell] = value
        sheet[cell].number_format = CURRENCY_FORMAT
    indicator_row = 28 + extra
    indicators = {
        (indicator_row, 5): value_total,
        (indicator_row, 11): team_total + freight_total,
        (indicator_row + 1, 5): ((team_total + freight_total) / value_total if value_total else 0),
        (indicator_row + 1, 11): value_total - team_total - freight_total,
        (indicator_row + 2, 5): round(
            team_total
            + freight_total
            - sum(float(work.get("custo_viagem") or 0) for work in works),
            2,
        ),
        (indicator_row + 2, 11): team_total + freight_total,
    }
    for (row, column), value in indicators.items():
        sheet.cell(row, column).value = value
        sheet.cell(row, column).number_format = (
            PERCENT_FORMAT if (row, column) == (indicator_row + 1, 5) else CURRENCY_FORMAT
        )
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_area = f"B2:M{36 + extra}"
    return extra


def _conversion_program() -> str:
    # LibreOffice é opcional e, quando presente, vem no Runtime assinado da aplicação.
    # Suporta PortableApps e a árvore convencional, sem procurar instalação da estação.
    master = runtime_libreoffice()
    return str(master) if master is not None else ""


def _valid_pdf(path: Path) -> bool:
    try:
        if path.stat().st_size < 1024:
            return False
        with path.open("rb") as stream:
            if stream.read(5) != b"%PDF-":
                return False
            stream.seek(max(0, path.stat().st_size - 2048))
            return b"%%EOF" in stream.read()
    except OSError:
        return False


def _run_pdf_conversion(command: list[str], expected: Path) -> None:
    process_options = {
        "stdout": subprocess.PIPE,
        "stderr": subprocess.PIPE,
        "text": True,
    }
    if os.name == "nt":
        process_options["creationflags"] = subprocess.CREATE_NEW_PROCESS_GROUP
    else:
        process_options["start_new_session"] = True
    process = subprocess.Popen(command, **process_options)
    deadline = time.monotonic() + 90
    stable_size = -1
    stable_since = 0.0
    while time.monotonic() < deadline:
        return_code = process.poll()
        if _valid_pdf(expected):
            size = expected.stat().st_size
            if size != stable_size:
                stable_size = size
                stable_since = time.monotonic()
            elif time.monotonic() - stable_since >= 1.0:
                break
        if return_code is not None:
            break
        time.sleep(0.2)

    if process.poll() is None:
        if os.name == "nt":
            subprocess.run(
                ["taskkill", "/PID", str(process.pid), "/T", "/F"],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, check=False,
            )
        else:
            try:
                os.killpg(process.pid, signal.SIGTERM)
            except OSError:
                process.terminate()
        try:
            process.wait(timeout=5)
        except subprocess.TimeoutExpired:
            if os.name != "nt":
                try:
                    os.killpg(process.pid, signal.SIGKILL)
                except OSError:
                    process.kill()
            else:
                process.kill()
            process.wait(timeout=5)
    stdout, stderr = process.communicate()
    if not _valid_pdf(expected):
        detail = (stderr or stdout or "TEMPO LIMITE OU PDF INVÁLIDO").strip()
        raise RuntimeError(detail)


def _convert_pdf(
    workbook_path: Path, sheets: list[str], pdf_dir: Path, temp_root: Path
) -> list[Path]:
    program = _conversion_program()
    if not program:
        return []
    resource_root = resource_target("LIBREOFFICE")
    try:
        Path(program).resolve().relative_to(resource_root.resolve())
        verify_installed_resource("LIBREOFFICE")
    except ValueError:
        # LibreOffice integrado ao Mestre/Runtime oficial: mantém a verificação
        # tradicional do componente assinado.
        verify_runtime_component(runtime_component_installation_root("LibreOffice"), "LibreOffice")
    load_workbook, _ = _require_openpyxl()
    outputs = []
    pdf_dir.mkdir(parents=True, exist_ok=True)
    temp_root.mkdir(parents=True, exist_ok=True)
    for sheet_name in sheets:
        profile = temp_root / f"LibreOfficeProfile-{sheet_name}"
        profile.mkdir()
        book = load_workbook(workbook_path)
        for name in list(book.sheetnames):
            if name != sheet_name:
                del book[name]
        single = temp_root / f"{sheet_name}.xlsx"
        book.save(single)
        book.close()
        try:
            _run_pdf_conversion(
                [
                    program,
                    f"-env:UserInstallation={profile.resolve().as_uri()}",
                    "--headless",
                    "--convert-to",
                    "pdf",
                    "--outdir",
                    str(pdf_dir),
                    str(single),
                ],
                pdf_dir / f"{sheet_name}.pdf",
            )
        except (OSError, subprocess.SubprocessError, RuntimeError) as exc:
            raise RuntimeError(f"FALHA AO GERAR O PDF {sheet_name}: {exc}") from exc
        expected = pdf_dir / f"{sheet_name}.pdf"
        outputs.append(expected)
    return outputs


def _public_document(row) -> dict:
    item = dict(row)
    item.pop("workbook_path", None)
    item["arquivos"] = json.loads(item.pop("arquivos_json") or "[]")
    item["arquivos"] = [file for file in item["arquivos"] if file.get("tipo") in {"EXCEL", "PDF"}]
    for file in item["arquivos"]:
        file["download_url"] = (
            f"/api/documentos/{item['id']}/arquivos/{quote(file['nome'], safe='')}"
        )
    return item


def list_documents(load_id: str) -> list[dict]:
    with connect() as connection:
        rows = connection.execute(
            "SELECT * FROM carregamento_documentos WHERE carregamento_id=? ORDER BY revisao DESC",
            (load_id,),
        )
        return [_public_document(row) for row in rows]


def capture_package_input(load_id: str) -> dict:
    """Captura, dentro de uma leitura sincronizada, a visão usada na geração.

    O trabalho pesado de Excel/PDF pode ocorrer depois que o lock é liberado;
    a publicação curta revalida a revisão e o hash do conteúdo.
    """
    if not TEMPLATE_PATH.is_file():
        raise RuntimeError("TEMPLATE PADRÃO DE CARREGAMENTO NÃO ENCONTRADO.")
    carregamentos.assert_can_modify(load_id, allow_expedited=True)
    original_load = carregamentos.get_carregamento(load_id)
    identity = current_identity()
    with connect() as connection:
        previous = connection.execute(
            "SELECT COALESCE(MAX(revisao),0) FROM carregamento_documentos WHERE carregamento_id=?",
            (load_id,),
        ).fetchone()[0]
    return {
        "format": 1,
        "load_id": load_id,
        "load": deepcopy(original_load),
        "previous_revision": int(previous),
        "operational_revision": int(original_load.get("revisao_operacional") or 1),
        "content_hash": _json_hash(_document_content(original_load)),
        "template_hash": _sha256_file(TEMPLATE_PATH),
        "identity": {
            "user_id": identity.user_id,
            "user_name": identity.user_name,
            "station_id": identity.station_id or station_id(),
        },
    }


def prepare_package(load_id: str, captured: dict | None = None) -> dict:
    captured = captured or capture_package_input(load_id)
    if int(captured.get("format") or 0) != 1 or captured.get("load_id") != load_id:
        raise RuntimeError("CAPTURA DOCUMENTAL INVÁLIDA.")
    original_load = deepcopy(captured.get("load") or {})
    if original_load.get("id") != load_id:
        raise RuntimeError("CAPTURA DOCUMENTAL NÃO CORRESPONDE AO CARREGAMENTO.")
    template_hash = str(captured.get("template_hash") or "")
    content_hash = str(captured.get("content_hash") or "")
    if len(template_hash) != 64 or len(content_hash) != 64:
        raise RuntimeError("HASH DA CAPTURA DOCUMENTAL INVÁLIDO.")
    identity = current_identity()
    expected_identity = captured.get("identity") or {}
    if (
        identity.user_id != expected_identity.get("user_id")
        or identity.user_name != expected_identity.get("user_name")
        or (identity.station_id or station_id()) != expected_identity.get("station_id")
    ):
        raise PermissionError("A IDENTIDADE MUDOU APÓS A CAPTURA DOCUMENTAL.")
    previous = int(captured.get("previous_revision") or 0)
    load = _excel_safe(original_load)
    revision = int(previous) + 1
    document_id = "PENDENTE"
    final_destination = _document_root(original_load, revision)
    if final_destination.exists():
        raise RuntimeError("A PASTA DA NOVA REVISÃO DOCUMENTAL JÁ EXISTE.")
    staging_root = _safe_output(DATA_ROOT / ".CJLStaging" / "Documentos")
    staging_root.mkdir(parents=True, exist_ok=True)
    for stale in staging_root.iterdir():
        try:
            if stale.is_dir() and time.time() - stale.stat().st_mtime > 24 * 60 * 60:
                shutil.rmtree(stale, ignore_errors=True)
        except OSError:
            pass
    prepared_root = Path(tempfile.mkdtemp(prefix=f"{load_id}-REV{revision:03d}-", dir=staging_root))
    destination = prepared_root / "Pacote"
    workbook_name = f"{load_id}__DOCUMENTOS.xlsx"
    generated_at = _now()
    files = []
    with tempfile.TemporaryDirectory(prefix="cjl-doc-", dir=prepared_root) as temp_name:
        temp_root = Path(temp_name)
        workbook_path = temp_root / workbook_name
        load_workbook, Image = _require_openpyxl()
        book = load_workbook(TEMPLATE_PATH)
        matrix = book["OBRA_01"]
        summary = book["RESUMO_VALORES"]
        logic = book["_LOGICA_SISTEMA"]
        work_sheets = [matrix]
        for index in range(2, len(load.get("obras", [])) + 1):
            sheet = book.copy_worksheet(matrix)
            sheet.title = f"OBRA_{index:02d}"
            book._sheets.remove(sheet)
            book._sheets.insert(index, sheet)
            work_sheets.append(sheet)
        for sheet, work in zip(work_sheets, load.get("obras", [])):
            _fill_work_sheet(
                sheet,
                load,
                work,
                _work_items(load, work),
                generated_at,
                content_hash,
                Image,
            )
        _write_cost_sheet(book["CARREGAMENTO_CUSTO"], load, generated_at)
        _write_summary_sheet(summary, load, generated_at)
        logic.sheet_state = "hidden"
        logic["B51"] = "METADADOS DA GERAÇÃO MAIS RECENTE"
        provenance = public_notice()
        metadata_rows = [
            ("VERSÃO", app_version()),
            ("REVISÃO", revision),
            ("TEMPLATE_SHA256", template_hash),
            ("CONTEÚDO_SHA256", content_hash),
            ("USUÁRIO", identity.user_name),
            ("ESTAÇÃO", identity.station_id or station_id()),
            ("TIMESTAMP", generated_at),
            ("CRIADOR_TITULAR", provenance["creator"]),
            ("IDENTIFICACAO_PUBLICA", provenance["public_id"]),
            ("LICENCA_SHA256", provenance["license_sha256"]),
            ("MESTRE", provenance["master_id"]),
        ]
        for offset, values in enumerate(metadata_rows, start=52):
            logic.cell(offset, 2).value, logic.cell(offset, 3).value = values
        if book.calculation is not None:
            book.calculation.fullCalcOnLoad = True
            book.calculation.forceFullCalc = True
        book.save(workbook_path)
        book.close()
        output_sheets = ["CARREGAMENTO_CUSTO"] + [
            f"OBRA_{index:02d}" for index in range(1, len(load.get("obras", [])) + 1)
        ] + ["RESUMO_VALORES"]
        pdf_warning = None
        try:
            pdf_files = _convert_pdf(
                workbook_path, output_sheets, temp_root / "PDF", temp_root / "Singles"
            )
        except Exception as exc:
            # PDF é um recurso adicional. Falha do LibreOffice nunca invalida o Excel já gerado.
            pdf_files = []
            pdf_warning = (
                "EXCEL GERADO NORMALMENTE. A EXPORTAÇÃO PDF OPCIONAL FALHOU: "
                f"{exc}"
            )
        if not pdf_files and not pdf_warning:
            candidates = " | ".join(str(path) for path in runtime_libreoffice_candidates())
            pdf_warning = (
                "EXCEL GERADO NORMALMENTE. PDFs NÃO FORAM GERADOS PORQUE O LIBREOFFICE OPCIONAL "
                f"NÃO FOI LOCALIZADO NO RUNTIME ASSINADO. CAMINHOS ACEITOS: {candidates}"
            )
        package = temp_root / "Pacote"
        package.mkdir()
        shutil.copy2(workbook_path, package / workbook_name)
        for pdf in pdf_files:
            shutil.copy2(pdf, package / pdf.name)
        for file in sorted(package.iterdir()):
            files.append(
                {
                    "nome": file.name,
                    "tipo": "EXCEL" if file.suffix.lower() == ".xlsx" else "PDF",
                    "tamanho": file.stat().st_size,
                    "sha256": _sha256_file(file),
                }
            )
        manifest = {
            "documento_id": document_id,
            "carregamento_id": load_id,
            "revisao": revision,
            "template_sha256": template_hash,
            "conteudo_sha256": content_hash,
            "usuario_id": identity.user_id,
            "usuario_nome": identity.user_name,
            "estacao_id": identity.station_id or station_id(),
            "gerado_em": generated_at,
            "pdf_status": "GERADO" if pdf_files else "INDISPONIVEL",
            "pdf_aviso": pdf_warning,
            "arquivos": files,
        }
        atomic_write_json(package / "manifesto_documental.json", manifest)
        files.append(
            {
                "nome": "manifesto_documental.json",
                "tipo": "MANIFESTO",
                "tamanho": (package / "manifesto_documental.json").stat().st_size,
                "sha256": _sha256_file(package / "manifesto_documental.json"),
            }
        )
        os.replace(package, destination)
    package_bytes = sum(file["tamanho"] for file in files)
    ensure_data_quota(DATA_ROOT, package_bytes)
    return {
        "format": 1,
        "load_id": load_id,
        "revision": revision,
        "previous_revision": int(previous),
        "operational_revision": int(captured.get("operational_revision") or 1),
        "content_hash": content_hash,
        "template_hash": template_hash,
        "workbook_name": workbook_name,
        "generated_at": generated_at,
        "identity": dict(expected_identity),
        "files": files,
        "package": str(destination),
        "prepared_root": str(prepared_root),
        "final_destination": str(final_destination),
        "pdf_warning": pdf_warning,
        "pdf_generated": bool(pdf_files),
    }


def discard_prepared(prepared: dict) -> None:
    root = Path(str(prepared.get("prepared_root") or ""))
    staging_root = _safe_output(DATA_ROOT / ".CJLStaging" / "Documentos")
    try:
        root.resolve().relative_to(staging_root.resolve())
    except (OSError, ValueError):
        return
    shutil.rmtree(root, ignore_errors=True)


def commit_prepared_package(prepared: dict) -> dict:
    if int(prepared.get("format") or 0) != 1:
        raise RuntimeError("PREPARAÇÃO DOCUMENTAL INVÁLIDA.")
    load_id = str(prepared.get("load_id") or "")
    carregamentos.assert_can_modify(load_id, allow_expedited=True)
    current_load = carregamentos.get_carregamento(load_id)
    if int(current_load.get("revisao_operacional") or 0) != int(prepared.get("operational_revision") or -1):
        raise RuntimeError("O CARREGAMENTO MUDOU DURANTE A GERAÇÃO. GERE OS DOCUMENTOS NOVAMENTE.")
    if _json_hash(_document_content(current_load)) != prepared.get("content_hash"):
        raise RuntimeError("O CONTEÚDO DO CARREGAMENTO MUDOU DURANTE A GERAÇÃO.")
    package = _safe_output(Path(str(prepared.get("package") or "")))
    destination = _safe_output(Path(str(prepared.get("final_destination") or "")))
    staging_root = _safe_output(DATA_ROOT / ".CJLStaging" / "Documentos")
    try:
        package.resolve().relative_to(staging_root.resolve())
    except ValueError as exc:
        raise RuntimeError("PACOTE PREPARADO FORA DA ÁREA DE ESTÁGIO.") from exc
    if not package.is_dir() or destination.exists():
        raise RuntimeError("PACOTE PREPARADO AUSENTE OU DESTINO DOCUMENTAL JÁ EXISTENTE.")
    identity = current_identity()
    expected_identity = prepared.get("identity") or {}
    if (
        identity.user_id != expected_identity.get("user_id")
        or identity.user_name != expected_identity.get("user_name")
        or (identity.station_id or station_id()) != expected_identity.get("station_id")
    ):
        raise PermissionError(
            "A IDENTIDADE DO USUÁRIO MUDOU DURANTE A GERAÇÃO. GERE O PACOTE NOVAMENTE."
        )
    revision = int(prepared["revision"])
    with connect() as connection:
        previous = int(connection.execute(
            "SELECT COALESCE(MAX(revisao),0) FROM carregamento_documentos WHERE carregamento_id=?",
            (load_id,),
        ).fetchone()[0])
        if previous != int(prepared["previous_revision"]) or revision != previous + 1:
            raise RuntimeError("OUTRA REVISÃO DOCUMENTAL FOI PUBLICADA. GERE NOVAMENTE.")
        document_id = next_id(connection, "documento", "DOC")
        manifest_path = package / "manifesto_documental.json"
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        manifest["documento_id"] = document_id
        atomic_write_json(manifest_path, manifest)
        files = [file for file in prepared["files"] if file.get("nome") != manifest_path.name]
        files.append({
            "nome": manifest_path.name,
            "tipo": "MANIFESTO",
            "tamanho": manifest_path.stat().st_size,
            "sha256": _sha256_file(manifest_path),
        })
        workbook_hash = next(file["sha256"] for file in files if file["tipo"] == "EXCEL")
        stage_directory(package, destination)
        connection.execute(
            """INSERT INTO carregamento_documentos(
                   id,carregamento_id,revisao,tipo,workbook_nome,workbook_path,workbook_sha256,
                   template_sha256,conteudo_sha256,arquivos_json,usuario_id,usuario_nome,
                   estacao_id,gerado_em
               ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                document_id, load_id, revision, "PACOTE", prepared["workbook_name"],
                (destination / prepared["workbook_name"]).relative_to(DATA_ROOT).as_posix(),
                workbook_hash, prepared["template_hash"], prepared["content_hash"],
                json.dumps(files, ensure_ascii=False, allow_nan=False),
                identity.user_id, identity.user_name, identity.station_id or station_id(),
                prepared["generated_at"],
            ),
        )
    result = next(row for row in list_documents(load_id) if row["id"] == document_id)
    if prepared.get("pdf_warning"):
        result["aviso"] = prepared["pdf_warning"]
    result["pdf_gerado"] = bool(prepared.get("pdf_generated"))
    return result


def generate_package(load_id: str) -> dict:
    if current_file_transaction() is None:
        raise RuntimeError("A PUBLICAÇÃO DOCUMENTAL EXIGE UMA TRANSAÇÃO OFICIAL ATIVA.")
    prepared = prepare_package(load_id)
    try:
        return commit_prepared_package(prepared)
    except Exception:
        discard_prepared(prepared)
        raise


def document_file(document_id: str, file_name: str) -> tuple[Path, dict, dict]:
    safe_name = Path(str(file_name or "").replace("\\", "/")).name
    with connect() as connection:
        row = connection.execute(
            "SELECT * FROM carregamento_documentos WHERE id=?", (document_id,)
        ).fetchone()
    if not row:
        raise FileNotFoundError("DOCUMENTO NÃO ENCONTRADO.")
    document = dict(row)
    document["arquivos"] = json.loads(document.pop("arquivos_json") or "[]")
    metadata = next(
        (file for file in document["arquivos"] if file["nome"] == safe_name), None
    )
    if not metadata:
        raise FileNotFoundError("ARQUIVO DOCUMENTAL NÃO ENCONTRADO.")
    base = _safe_output(DATA_ROOT / Path(row["workbook_path"]).parent)
    path = _safe_output(base / safe_name)
    if not path.is_file() or _sha256_file(path) != metadata["sha256"]:
        raise RuntimeError("O HASH DO ARQUIVO DOCUMENTAL NÃO CONFERE; DOWNLOAD BLOQUEADO.")
    mime = {
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".pdf": "application/pdf",
        ".json": "application/json",
    }.get(path.suffix.lower(), "application/octet-stream")
    return path, metadata, {
        "mime": mime,
        "nome_original": safe_name,
        "nome_seguro": safe_name,
    }


def validate_documents(load_id: str) -> dict:
    failures = []
    with connect() as connection:
        documents = []
        for row in connection.execute("SELECT * FROM carregamento_documentos WHERE carregamento_id=?", (load_id,)):
            item = dict(row)
            item["arquivos"] = json.loads(item.pop("arquivos_json") or "[]")
            documents.append(item)
    for document in documents:
        for metadata in document["arquivos"]:
            try:
                document_file(document["id"], metadata["nome"])
            except Exception as exc:
                failures.append(
                    {
                        "documento": document["id"],
                        "arquivo": metadata["nome"],
                        "erro": str(exc),
                    }
                )
    return {"ok": not failures, "documentos": len(documents), "falhas": failures}


def render_work_export(carregamento_id: str, obra_id: str) -> str:
    """Compatibilidade com o link HTML anterior à V1.008."""
    load = carregamentos.get_carregamento(carregamento_id)
    work = next((row for row in load.get("obras", []) if row["id"] == obra_id), None)
    if not work:
        raise ValueError("A OBRA INFORMADA NÃO PERTENCE AO CARREGAMENTO.")
    return (
        "<!doctype html><meta charset='utf-8'><title>CJL System</title>"
        "<style>body{font-family:Arial;padding:40px;color:#17344d;line-height:1.5}</style>"
        f"<h1>{html.escape(str(load['id']))} · {html.escape(str(work['nome']))}</h1>"
        "<p>A exportação oficial agora é gerada em Excel e PDF pelo botão "
        "<b>GERAR DOCUMENTOS</b> no carregamento.</p>"
    )
